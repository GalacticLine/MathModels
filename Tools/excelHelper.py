""" Excel导出辅助工具 """
import time
import warnings
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment
import pandas as pd


def permission_sleep(func):
    """
    权限错误等待装饰器，
    保证程序出现权限错误时不会中断，程序将等待短暂时间后再次尝试执行函数。
    :param func: 要装饰的函数
    :return:
    """

    def wrapper(*args, **kwargs):
        while True:
            try:
                result = func(*args, **kwargs)
                break
            except PermissionError as e:
                warnings.warn(f'{e} 程序将等待权限错误解决后，继续运行。')
                time.sleep(1)
        return result
    return wrapper


def set_three_line(path: str, start_row=1):
    """
    设置三线表风格
    :param start_row: 首行，默认为第一行
    :param path: excel文件路径
    :return:
    """
    wb = load_workbook(filename=path)
    sheet = wb.active
    row_cells = sheet[start_row]
    border = Border(top=Side(style='thick'), bottom=Side(style='thin'))

    for cell in row_cells:
        cell.border = border

    row_cells = sheet[sheet.max_row]
    border = Border(bottom=Side(style='thick'))

    for cell in row_cells:
        cell.border = border
    wb.save(path)
    wb.close()


def marge_excel_col(xlsx, col=1):
    """
    合并Excel表格中指定列中相同数值的单元格，并居中对齐

    :param xlsx: Excel文件路径
    :param col: 要合并的列，默认为第1列
    :return:
    """
    try:
        wb = load_workbook(xlsx)
        sheet = wb.active
        unique_values = {}
        for row in range(2, sheet.max_row + 1):
            cell_value = sheet.cell(row=row, column=col).value
            if cell_value not in unique_values:
                unique_values[cell_value] = {'start_row': row, 'end_row': row}
            else:
                unique_values[cell_value]['end_row'] = row
        for value, rows in unique_values.items():
            start_row = rows['start_row']
            end_row = rows['end_row']
            sheet.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)
            merge_range = sheet.cell(row=start_row, column=col).coordinate + ":"
            merge_range += sheet.cell(row=end_row, column=col).coordinate
            sheet[merge_range][0][0].alignment = Alignment(horizontal='center', vertical='center')

        wb.save(xlsx)
        wb.close()
    except Exception as e:
        print(f"未知错误：{e}!")


def read_excel_first(path: str, sheet_name: str | int = 0):
    """
    从excel文件中读取首个数据,按空行分隔
    :param path: excel文件路径
    :param sheet_name: 工作表名
    :return: 层结构实例
    """
    df = pd.read_excel(path, sheet_name)
    empty_row = df[df.isnull().all(axis=1)].index
    if len(empty_row) > 0:
        df = df.head(empty_row[0])
    new_df = df.set_index(df.columns[0])
    new_df.index.name = None
    return new_df
